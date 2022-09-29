from numpy.core.numeric import full_like
import xlrd
import pyodbc
import urllib
import time
import pandas as pd
from pipeline_Plant_Trial.prepare import from_sheet_to_df,interval_to_string,get_abbreviation,valuecolumn_to_value_string, get_df_string_from_df,from_sheet_to_df_mutlipetime
import operator
from openpyxl import load_workbook

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def readtest():
    filepath_excel="CL_Jartest_Data/Cl Jar test combined_template.xlsx"
    calculate_columns_sheetname='CalculatedColumns(RE)'
    setup_sheetname='Setup(RE)'
    sheetnames=get_sheetnames_xlsx(filepath_excel)
    

    print (sheetnames) 
    sheet_df_list=[]
    header_dictionary={}

    #get setup parameters from setup sheet
    setup_df=pd.read_excel(filepath_excel,sheet_name=setup_sheetname)
    setup_condition_columns=[]
    setup_product_columns=[]
    setup_product_dosage_columns=[]
    setup_identifier_columns=['sheetname']   #sheet name as identifier always
    KPI_columns=[]


    for idx,row in setup_df.iterrows(): 
        if row['key column setup']=='condition':
            setup_condition_columns=str(row['column names']).split("$")
        if row['key column setup']=='KPI':
            KPI_columns=str(row['column names']).split("$")
        if row['key column setup']=='product':
            setup_product_columns=str(row['column names']).split("$")
        if row['key column setup']=='product dosage':
            setup_product_dosage_columns=str(row['column names']).split("$")
        if row['key column setup']=='identifier':
            setup_identifier_columns=str(row['column names']).split("$")

    # for idx,row in setup_df.iterrows():  

    #     if row['RequiredColumns/Parameters']=='Time_interval':
    #         time_interval=row['Excel_Column/Input']
    #     if row['RequiredColumns/Parameters']=='Dosage_interval_count':
    #         try:
    #             dosage_interval_count=int(row['Excel_Column/Input'])+1
    #         except:
    #             dosage_interval_count=5


    # if Interval_columns_str:
    #     Interval_columns=Interval_columns_str.split("$")
    # if KPI_columns_str:
    #     KPI_columns=KPI_columns_str.split("$")

    keyword_datetime='Date'


    sheet_df_list=[]
    header_dictionary={}
    for sheetname in sheetnames:
        if "RE" not in sheetname:
            sheetdf,sheetdf_headdic=from_sheet_to_df( excelname=filepath_excel,
                                                sheetname=sheetname,
                                                skip_rows_count=1,
                                                header_rows_count=2,
                                                keyword_datetime=keyword_datetime
                                                )
            
            # sheetdf2=from_sheet_to_df_mutlipetime( excelname=filepath_excel,
            #                                     sheetname=sheetname,
            #                                     skip_rows_count=1,
            #                                     header_rows_count=2,
            #                                     keyword_datetime=keyword_datetime
            #                                     )



            sheetdf = sheetdf.fillna(method='ffill')

            ###add sheet name as column
            
            sheetdf.insert(0,'sheetname',sheetname)

            sheet_df_list.append(sheetdf)
            a=1
    # concat dataframe 
    concated_df=pd.concat(sheet_df_list)

    ###
    setup_product_dosage_dic={}

    #assume dosage is associated with left column
    for columnname in concated_df.columns:
        if columnname in setup_product_columns:
            previous_product_column_name=columnname
        if columnname in setup_product_dosage_columns or (columnname.split("_")[0] in setup_product_dosage_columns):
            completedosagecolumnname=previous_product_column_name+'_'+setup_product_dosage_columns[0]
            setup_product_dosage_dic[previous_product_column_name]=completedosagecolumnname
            concated_df.rename({columnname: completedosagecolumnname}, axis=1, inplace=True)

    chemcial_dosage_list=[]
    for key in setup_product_dosage_dic.keys():
        chemcial_dosage_list.append(key)
        chemcial_dosage_list.append(setup_product_dosage_dic[key])

    defaultshow_columns=setup_identifier_columns+chemcial_dosage_list+KPI_columns+setup_condition_columns

    #process identifiier column
    #0 is sheet, so 1 is the primary identifier in each sheet
    concated_df[setup_identifier_columns[1]] = concated_df["sheetname"].astype(str) +'_'+ concated_df[setup_identifier_columns[1]].astype(str)
    





    #########################wirte to excel: debug only##################
    writer = pd.ExcelWriter('DashExport'+'/'+'Dashoutput.xlsx')
    #concated_df.to_excel(writer, sheet_name='test')
    concated_df.to_csv('DashExport'+'/'+'Dashout.csv')
    # sheet_merged_interval_string_df.to_csv('DashExport'+'/'+'Dashout_st_Interval.csv')
    # sheet_int_unpivotdf.to_csv('DashExport'+'/'+'finalUnpivot_int.csv')
    ########################export debug finished#########################

    return concated_df, defaultshow_columns,setup_identifier_columns, setup_condition_columns,KPI_columns, setup_product_dosage_dic




dataa=readtest()


#combine dataframe vertically
#https://statisticsglobe.com/combine-pandas-dataframes-vertically-horizontally-python

a=1